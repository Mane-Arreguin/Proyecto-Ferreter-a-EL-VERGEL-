import os
import sys
import subprocess
import unicodedata
from copy import copy
from datetime import datetime, timedelta
import pyodbc
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
SERVIDOR = r"DESKTOP-6UH4INV"
BASE_DATOS = "Ferreteria_ElVergel"
ARCHIVO_RUTA_EXCEL = "ruta_excel_el_vergel.txt"
LOGO_EMPRESA = "logo_el_vergel.png"
ARCHIVO_POWER_BI = "Reportes_Ferreteria el Vergel.pbix"
DATASET_VENTAS_2024 = "Dataset_Ferreteria_El_Vergel_2024.xlsx"
DATASET_VENTAS_2025 = "Dataset_Ferreteria_El_Vergel_2025.xlsx"
RUTA_EXCEL_ACTUAL = ""
COLOR_AZUL_OSCURO = "#002B5C"
COLOR_AZUL = "#004AAD"
COLOR_AZUL_CLARO = "#EAF2FF"
COLOR_BLANCO = "#FFFFFF"
COLOR_FONDO = "#F3F6FA"
COLOR_TEXTO = "#0B1F3A"
COLOR_GRIS = "#D9E2EC"
COLOR_VERDE = "#16A34A"
COLOR_NARANJA = "#F59E0B"
def obtener_driver():
    drivers = pyodbc.drivers()
    if "ODBC Driver 18 for SQL Server" in drivers:
        return "ODBC Driver 18 for SQL Server"
    if "ODBC Driver 17 for SQL Server" in drivers:
        return "ODBC Driver 17 for SQL Server"
    return "SQL Server"
def cadena_conexion(nombre_base):
    driver = obtener_driver()
    return (
        f"DRIVER={{{driver}}};"
        f"SERVER={SERVIDOR};"
        f"DATABASE={nombre_base};"
        "Trusted_Connection=yes;"
        "TrustServerCertificate=yes;"
        "MARS_Connection=yes;"
    )
def conectar_base():
    return pyodbc.connect(cadena_conexion(BASE_DATOS))
def conectar_master():
    return pyodbc.connect(cadena_conexion("master"), autocommit=True)
def crear_base_si_no_existe():
    con = conectar_master()
    cur = con.cursor()
    cur.execute("SELECT DB_ID(?) AS id_base", BASE_DATOS)
    fila = cur.fetchone()
    if fila.id_base is None:
        cur.execute(f"CREATE DATABASE [{BASE_DATOS}]")
        print("Base de datos creada:", BASE_DATOS)
    else:
        print("Base de datos ya existe:", BASE_DATOS)
    cur.close()
    con.close()
def ejecutar_bloques(cursor, bloques):
    for bloque in bloques:
        cursor.execute(bloque)
def preparar_sql_server():
    print("\nPreparando SQL Server...")
    print("Advertencia: esto recrea tablas y borra ventas previas.")
    crear_base_si_no_existe()
    con = conectar_base()
    cur = con.cursor()
    ejecutar_bloques(cur, [
        "IF OBJECT_ID('Ventas_2026','U') IS NOT NULL DROP TABLE Ventas_2026",
        "IF OBJECT_ID('Control_Stock','U') IS NOT NULL DROP TABLE Control_Stock",
        "IF OBJECT_ID('Catalogo_Ampliado','U') IS NOT NULL DROP TABLE Catalogo_Ampliado",
        "IF OBJECT_ID('Usuarios','U') IS NOT NULL DROP TABLE Usuarios",
        """
        CREATE TABLE Usuarios (
            id_usuario INT IDENTITY(1,1) PRIMARY KEY,
            usuario NVARCHAR(50) NOT NULL UNIQUE,
            password_hash VARBINARY(32) NOT NULL,
            password_salt VARBINARY(16) NOT NULL,
            rol NVARCHAR(30) NOT NULL,
            activo BIT NOT NULL DEFAULT 1
        )
        """,
        """
        CREATE TABLE Catalogo_Ampliado (
            sku NVARCHAR(30) PRIMARY KEY,
            categoria NVARCHAR(100),
            producto NVARCHAR(150),
            unidad NVARCHAR(50),
            costo_promedio DECIMAL(12,2),
            precio_venta DECIMAL(12,2),
            margen DECIMAL(10,4),
            iva_incluido NVARCHAR(20),
            proveedor NVARCHAR(100),
            lead_time_dias INT,
            stock_minimo DECIMAL(12,2),
            punto_reorden DECIMAL(12,2),
            stock_maximo DECIMAL(12,2),
            stock_actual DECIMAL(12,2),
            estado_stock NVARCHAR(30),
            ubicacion NVARCHAR(100),
            fecha_alta DATE,
            observacion NVARCHAR(255)
        )
        """,
        """
        CREATE TABLE Control_Stock (
            sku NVARCHAR(30) PRIMARY KEY,
            categoria NVARCHAR(100),
            producto NVARCHAR(150),
            unidad NVARCHAR(50),
            stock_inicial DECIMAL(12,2),
            entradas_compras DECIMAL(12,2),
            salidas_ventas DECIMAL(12,2),
            merma_ajuste DECIMAL(12,2),
            stock_calculado DECIMAL(12,2),
            stock_fisico DECIMAL(12,2),
            diferencia DECIMAL(12,2),
            stock_minimo DECIMAL(12,2),
            punto_reorden DECIMAL(12,2),
            stock_maximo DECIMAL(12,2),
            estado NVARCHAR(30),
            ultima_reposicion DATE,
            proveedor NVARCHAR(100),
            observacion NVARCHAR(255)
        )
        """,
        """
        CREATE TABLE Ventas_2026 (
            id_venta NVARCHAR(30) PRIMARY KEY,
            fecha DATE,
            mes NVARCHAR(20),
            dia_semana NVARCHAR(20),
            categoria NVARCHAR(100),
            sku NVARCHAR(30),
            producto NVARCHAR(150),
            unidad NVARCHAR(50),
            cantidad DECIMAL(12,2),
            precio_unitario DECIMAL(12,2),
            total_mxn DECIMAL(12,2),
            metodo_pago NVARCHAR(50),
            tipo_cliente NVARCHAR(80),
            vendedor NVARCHAR(80),
            tipo_registro NVARCHAR(80),
            observacion NVARCHAR(255)
        )
        """,
        """
        CREATE OR ALTER PROCEDURE sp_crear_usuario
            @usuario NVARCHAR(50),
            @contrasena NVARCHAR(100),
            @rol NVARCHAR(30)
        AS
        BEGIN
            IF NOT EXISTS (SELECT 1 FROM Usuarios WHERE usuario = @usuario)
            BEGIN
                DECLARE @salt VARBINARY(16)
                DECLARE @hash VARBINARY(32)
                SET @salt = CRYPT_GEN_RANDOM(16)
                SET @hash = HASHBYTES('SHA2_256', @salt + CONVERT(VARBINARY(MAX), @contrasena))
                INSERT INTO Usuarios(usuario,password_hash,password_salt,rol)
                VALUES(@usuario,@hash,@salt,@rol)
            END
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_login
            @usuario NVARCHAR(50),
            @contrasena NVARCHAR(100)
        AS
        BEGIN
            SELECT id_usuario, usuario, rol
            FROM Usuarios
            WHERE usuario = @usuario
              AND activo = 1
              AND password_hash = HASHBYTES('SHA2_256', password_salt + CONVERT(VARBINARY(MAX), @contrasena))
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_resumen_general
        AS
        BEGIN
            SELECT
                COUNT(*) AS registros,
                SUM(total_mxn) AS venta_total,
                AVG(total_mxn) AS promedio_registro,
                MIN(fecha) AS fecha_inicio,
                MAX(fecha) AS fecha_corte,
                COUNT(DISTINCT sku) AS productos_vendidos
            FROM Ventas_2026
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_ventas_por_mes
        AS
        BEGIN
            SELECT
                MONTH(fecha) AS numero_mes,
                mes,
                COUNT(*) AS registros,
                SUM(total_mxn) AS venta_total,
                AVG(total_mxn) AS promedio
            FROM Ventas_2026
            GROUP BY MONTH(fecha), mes
            ORDER BY MONTH(fecha)
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_ventas_por_categoria
        AS
        BEGIN
            SELECT categoria, COUNT(*) AS registros, SUM(total_mxn) AS venta_total
            FROM Ventas_2026
            GROUP BY categoria
            ORDER BY venta_total DESC
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_buscar_ventas
            @producto NVARCHAR(100) = '',
            @categoria NVARCHAR(100) = ''
        AS
        BEGIN
            SELECT TOP 300
                id_venta, fecha, categoria, sku, producto, cantidad,
                precio_unitario, total_mxn, metodo_pago, tipo_cliente,
                vendedor, tipo_registro
            FROM Ventas_2026
            WHERE producto LIKE '%' + @producto + '%'
              AND categoria LIKE '%' + @categoria + '%'
            ORDER BY fecha DESC
        END
        """,
        """
        CREATE OR ALTER PROCEDURE sp_consultar_stock
            @estado NVARCHAR(30) = ''
        AS
        BEGIN
            SELECT
                sku, producto, categoria, unidad, stock_fisico, stock_minimo,
                punto_reorden, estado, proveedor, observacion
            FROM Control_Stock
            WHERE estado LIKE '%' + @estado + '%'
            ORDER BY
                CASE estado
                    WHEN 'Crítico' THEN 1
                    WHEN 'Reordenar' THEN 2
                    WHEN 'Vigilar' THEN 3
                    ELSE 4
                END
        END
        """
    ])
    cur.execute("EXEC sp_crear_usuario @usuario=?, @contrasena=?, @rol=?", "admin", "1234", "Administrador")
    con.commit()
    cur.close()
    con.close()
    print("\nSQL Server preparado correctamente.")
    print("Usuario: admin")
    print("Contraseña: 1234")
def limpiar_texto(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFD", texto).encode("ascii", "ignore").decode("utf-8")
    texto = texto.replace("_", " ").replace("/", " ").replace("%", "")
    return " ".join(texto.split())
def crear_mapa_columnas(hoja):
    mapa = {}
    for celda in hoja[1]:
        nombre = limpiar_texto(celda.value)
        if nombre:
            mapa[nombre] = celda.column
    return mapa
def obtener(fila, mapa, nombres, default=None):
    for nombre in nombres:
        limpio = limpiar_texto(nombre)
        if limpio in mapa:
            return fila[mapa[limpio] - 1]
    return default
def convertir_fecha(valor):
    if valor is None or valor == "":
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, int) or isinstance(valor, float):
        return (datetime(1899, 12, 30) + timedelta(days=valor)).date()
    if isinstance(valor, str):
        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"]:
            try:
                return datetime.strptime(valor.strip(), fmt).date()
            except:
                pass
    return valor
def convertir_fecha_formulario(valor):
    return convertir_fecha(valor)
def convertir_numero(valor):
    if valor is None or valor == "":
        return 0
    try:
        return float(valor)
    except:
        return 0
def convertir_entero(valor):
    if valor is None or valor == "":
        return 0
    try:
        return int(valor)
    except:
        return 0
def seleccionar_archivo_excel():
    ventana = tk.Tk()
    ventana.withdraw()
    ventana.attributes("-topmost", True)
    ruta = filedialog.askopenfilename(
        title="Selecciona el archivo Excel de El Vergel",
        initialdir=r"C:\Users\Manuel Arreguin\Downloads",
        filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
    )
    ventana.destroy()
    return ruta
def guardar_ruta_excel_actual(ruta):
    global RUTA_EXCEL_ACTUAL
    RUTA_EXCEL_ACTUAL = ruta
    try:
        with open(ARCHIVO_RUTA_EXCEL, "w", encoding="utf-8") as archivo:
            archivo.write(ruta)
    except:
        pass
def cargar_ruta_excel_actual():
    global RUTA_EXCEL_ACTUAL
    if RUTA_EXCEL_ACTUAL and os.path.exists(RUTA_EXCEL_ACTUAL):
        return RUTA_EXCEL_ACTUAL
    try:
        if os.path.exists(ARCHIVO_RUTA_EXCEL):
            ruta = open(ARCHIVO_RUTA_EXCEL, "r", encoding="utf-8").read().strip()
            if ruta and os.path.exists(ruta):
                RUTA_EXCEL_ACTUAL = ruta
                return ruta
    except:
        pass
    return ""
def obtener_ruta_excel_para_actualizar():
    ruta = cargar_ruta_excel_actual()
    if ruta:
        return ruta
    messagebox.showinfo("Seleccionar Excel", "Selecciona el archivo Excel que se va a actualizar.")
    ruta = seleccionar_archivo_excel()
    if ruta:
        guardar_ruta_excel_actual(ruta)
    return ruta
def limpiar_tablas(cursor):
    print("Limpiando tablas...")
    cursor.execute("DELETE FROM Ventas_2026")
    cursor.execute("DELETE FROM Control_Stock")
    cursor.execute("DELETE FROM Catalogo_Ampliado")
def importar_catalogo(cursor, libro):
    print("Importando Catalogo_Ampliado...")
    if "Catalogo_Ampliado" not in libro.sheetnames:
        print("No existe la hoja Catalogo_Ampliado")
        return
    hoja = libro["Catalogo_Ampliado"]
    mapa = crear_mapa_columnas(hoja)
    contador = 0
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        sku = obtener(fila, mapa, ["SKU"])
        if sku is None:
            continue
        cursor.execute("""
            INSERT INTO Catalogo_Ampliado (
                sku,categoria,producto,unidad,costo_promedio,precio_venta,margen,iva_incluido,
                proveedor,lead_time_dias,stock_minimo,punto_reorden,stock_maximo,stock_actual,
                estado_stock,ubicacion,fecha_alta,observacion
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        sku,
        obtener(fila, mapa, ["Categoría", "Categoria"]),
        obtener(fila, mapa, ["Producto"]),
        obtener(fila, mapa, ["Unidad"]),
        convertir_numero(obtener(fila, mapa, ["Costo Promedio MXN", "Costo Promedio"])),
        convertir_numero(obtener(fila, mapa, ["Precio Venta MXN", "Precio Venta"])),
        convertir_numero(obtener(fila, mapa, ["Margen", "Margen %"])),
        obtener(fila, mapa, ["IVA Incluido"]),
        obtener(fila, mapa, ["Proveedor"]),
        convertir_entero(obtener(fila, mapa, ["Lead Time Días", "Lead Time Dias", "Plazo Entrega Dias"])),
        convertir_numero(obtener(fila, mapa, ["Stock Mínimo", "Stock Minimo"])),
        convertir_numero(obtener(fila, mapa, ["Punto Reorden"])),
        convertir_numero(obtener(fila, mapa, ["Stock Máximo", "Stock Maximo"])),
        convertir_numero(obtener(fila, mapa, ["Stock Actual"])),
        obtener(fila, mapa, ["Estado Stock"], ""),
        obtener(fila, mapa, ["Ubicación", "Ubicacion"]),
        convertir_fecha(obtener(fila, mapa, ["Fecha Alta"])),
        obtener(fila, mapa, ["Observación Catálogo", "Observacion Catalogo", "Observación", "Observacion"], ""))
        contador += 1
    print("Catalogo importado:", contador)
def importar_stock(cursor, libro):
    print("Importando Control_Stock...")
    if "Control_Stock" not in libro.sheetnames:
        print("No existe la hoja Control_Stock")
        return
    hoja = libro["Control_Stock"]
    mapa = crear_mapa_columnas(hoja)
    contador = 0
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        sku = obtener(fila, mapa, ["SKU"])
        if sku is None:
            continue
        stock_inicial = convertir_numero(obtener(fila, mapa, ["Stock Inicial 2026", "Stock Inicial"]))
        entradas = convertir_numero(obtener(fila, mapa, ["Entradas/Compras 2026", "Entradas Compras 2026", "Entradas Compras"]))
        salidas = convertir_numero(obtener(fila, mapa, ["Salidas por Ventas", "Salidas Ventas"]))
        merma = convertir_numero(obtener(fila, mapa, ["Merma/Ajuste", "Merma Ajuste"], 0))
        stock_calculado = convertir_numero(obtener(fila, mapa, ["Stock Calculado"], stock_inicial + entradas - salidas - merma))
        stock_fisico = convertir_numero(obtener(fila, mapa, ["Stock Físico Conteo", "Stock Fisico Conteo", "Stock Físico", "Stock Fisico"]))
        diferencia = convertir_numero(obtener(fila, mapa, ["Diferencia"], stock_fisico - stock_calculado))
        cursor.execute("""
            INSERT INTO Control_Stock (
                sku,categoria,producto,unidad,stock_inicial,entradas_compras,salidas_ventas,
                merma_ajuste,stock_calculado,stock_fisico,diferencia,stock_minimo,
                punto_reorden,stock_maximo,estado,ultima_reposicion,proveedor,observacion
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        sku,
        obtener(fila, mapa, ["Categoría", "Categoria"]),
        obtener(fila, mapa, ["Producto"]),
        obtener(fila, mapa, ["Unidad"]),
        stock_inicial, entradas, salidas, merma, stock_calculado, stock_fisico, diferencia,
        convertir_numero(obtener(fila, mapa, ["Stock Mínimo", "Stock Minimo"])),
        convertir_numero(obtener(fila, mapa, ["Punto Reorden"])),
        convertir_numero(obtener(fila, mapa, ["Stock Máximo", "Stock Maximo"])),
        obtener(fila, mapa, ["Estado"], ""),
        convertir_fecha(obtener(fila, mapa, ["Última Reposición", "Ultima Reposicion"])),
        obtener(fila, mapa, ["Proveedor"]),
        obtener(fila, mapa, ["Observación Stock", "Observacion Stock", "Observación", "Observacion"], ""))
        contador += 1
    print("Stock importado:", contador)
def importar_ventas(cursor, libro):
    print("Importando Ventas_2026...")
    if "Ventas_2026" not in libro.sheetnames:
        print("No existe la hoja Ventas_2026")
        return
    hoja = libro["Ventas_2026"]
    mapa = crear_mapa_columnas(hoja)
    contador = 0
    for fila in hoja.iter_rows(min_row=2, values_only=True):
        id_venta = obtener(fila, mapa, ["ID_Venta", "ID Venta"])
        if id_venta is None:
            continue
        cursor.execute("""
            INSERT INTO Ventas_2026 (
                id_venta,fecha,mes,dia_semana,categoria,sku,producto,unidad,cantidad,
                precio_unitario,total_mxn,metodo_pago,tipo_cliente,vendedor,tipo_registro,observacion
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        id_venta,
        convertir_fecha(obtener(fila, mapa, ["Fecha"])),
        obtener(fila, mapa, ["Mes"]),
        obtener(fila, mapa, ["Dia_Semana", "Día Semana", "Dia Semana"]),
        obtener(fila, mapa, ["Categoría", "Categoria"]),
        obtener(fila, mapa, ["SKU"]),
        obtener(fila, mapa, ["Producto"]),
        obtener(fila, mapa, ["Unidad"]),
        convertir_numero(obtener(fila, mapa, ["Cantidad"])),
        convertir_numero(obtener(fila, mapa, ["Precio Unitario MXN", "Precio Unitario"])),
        convertir_numero(obtener(fila, mapa, ["Total MXN", "Total"])),
        obtener(fila, mapa, ["Metodo Pago", "Método Pago"]),
        obtener(fila, mapa, ["Tipo Cliente"]),
        obtener(fila, mapa, ["Vendedor"]),
        obtener(fila, mapa, ["Tipo Registro"]),
        obtener(fila, mapa, ["Observacion", "Observación"], ""))
        contador += 1
        if contador % 1000 == 0:
            print("Ventas importadas:", contador)
    print("Ventas importadas:", contador)
def validar_importacion(cursor):
    print("\nValidando importacion...")
    consultas = [
        ("Registros en ventas", "SELECT COUNT(*) AS valor FROM Ventas_2026"),
        ("Venta total", "SELECT SUM(total_mxn) AS valor FROM Ventas_2026"),
        ("Productos en catalogo", "SELECT COUNT(*) AS valor FROM Catalogo_Ampliado"),
        ("Registros en stock", "SELECT COUNT(*) AS valor FROM Control_Stock")
    ]
    for etiqueta, consulta in consultas:
        cursor.execute(consulta)
        print(etiqueta + ":", cursor.fetchone().valor)
    cursor.execute("SELECT MIN(fecha) AS inicio, MAX(fecha) AS corte FROM Ventas_2026")
    fechas = cursor.fetchone()
    print("Fecha inicio:", fechas.inicio)
    print("Fecha corte:", fechas.corte)
def importar_excel():
    ruta_excel = seleccionar_archivo_excel()
    if not ruta_excel:
        print("No seleccionaste ningun archivo.")
        return
    print("\nArchivo seleccionado:")
    print(ruta_excel)
    guardar_ruta_excel_actual(ruta_excel)
    con = conectar_base()
    cur = con.cursor()
    print("Leyendo Excel...")
    libro = load_workbook(ruta_excel, data_only=True)
    print("Hojas encontradas:")
    for hoja in libro.sheetnames:
        print("-", hoja)
    limpiar_tablas(cur)
    importar_catalogo(cur, libro)
    importar_stock(cur, libro)
    importar_ventas(cur, libro)
    con.commit()
    validar_importacion(cur)
    cur.close()
    con.close()
    print("\nImportacion terminada correctamente.")
def configurar_estilos():
    estilo = ttk.Style()
    try:
        estilo.theme_use("clam")
    except:
        pass
    estilo.configure("Treeview", background=COLOR_BLANCO, foreground=COLOR_TEXTO, rowheight=28,
                     fieldbackground=COLOR_BLANCO, font=("Segoe UI", 9))
    estilo.configure("Treeview.Heading", background=COLOR_AZUL_OSCURO, foreground=COLOR_BLANCO,
                     font=("Segoe UI", 9, "bold"))
    estilo.map("Treeview", background=[("selected", COLOR_AZUL)], foreground=[("selected", COLOR_BLANCO)])
    estilo.configure("TCombobox", padding=5, font=("Segoe UI", 10))
def cargar_logo_tk():
    try:
        img = tk.PhotoImage(file=LOGO_EMPRESA)
        if img.width() > 280:
            img = img.subsample(max(1, int(img.width() / 260)), max(1, int(img.width() / 260)))
        return img
    except:
        return None
def ruta_recurso(nombre):
    try:
        base = os.path.dirname(os.path.abspath(__file__))
    except:
        base = os.getcwd()
    return os.path.join(base, nombre)
def abrir_archivo_externo(nombre, titulo, tipos):
    ruta = ruta_recurso(nombre)
    if not os.path.exists(ruta):
        messagebox.showinfo("Seleccionar archivo", "No se encontro el archivo:\n\n" + nombre + "\n\nSelecciona su ubicacion manualmente.")
        ruta = filedialog.askopenfilename(title=titulo, initialdir=os.getcwd(), filetypes=tipos)
        if not ruta:
            return
    try:
        if os.name == "nt":
            os.startfile(ruta)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", ruta])
        else:
            subprocess.Popen(["xdg-open", ruta])
    except Exception as error:
        messagebox.showerror("No se pudo abrir el archivo", str(error))
def abrir_power_bi():
    abrir_archivo_externo(ARCHIVO_POWER_BI, "Selecciona el archivo de Power BI", [("Power BI", "*.pbix"), ("Todos los archivos", "*.*")])
def abrir_dataset_2024():
    abrir_archivo_externo(DATASET_VENTAS_2024, "Selecciona el dataset 2024", [("Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
def abrir_dataset_2025():
    abrir_archivo_externo(DATASET_VENTAS_2025, "Selecciona el dataset 2025", [("Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
def cargar_dataset_en_tabla(tabla, nombre_archivo):
    ruta = ruta_recurso(nombre_archivo)
    if not os.path.exists(ruta):
        ruta = filedialog.askopenfilename(title="Selecciona " + nombre_archivo, initialdir=os.getcwd(), filetypes=[("Excel", "*.xlsx"), ("Todos los archivos", "*.*")])
        if not ruta:
            return
    try:
        libro = load_workbook(ruta, data_only=True, read_only=True)
        hoja = libro[libro.sheetnames[0]]
        filas = list(hoja.iter_rows(values_only=True))
        if not filas:
            messagebox.showinfo("Dataset vacio", "El archivo no contiene filas.")
            return
        columnas = []
        for i, valor in enumerate(filas[0][:18], start=1):
            nombre = str(valor).strip() if valor is not None else "Columna " + str(i)
            if nombre in columnas:
                nombre = nombre + " " + str(i)
            columnas.append(nombre)
        datos = []
        for fila in filas[1:401]:
            datos.append(tuple("" if v is None else v for v in fila[:len(columnas)]))
        mostrar_datos(tabla, columnas, datos)
        messagebox.showinfo("Dataset cargado", "Se muestran hasta 400 registros de:\n\n" + os.path.basename(ruta) + "\n\nEste modulo no usa SQL Server.")
    except Exception as error:
        messagebox.showerror("Error al cargar dataset", str(error))
def boton_menu(panel, texto, comando=None, activo=False):
    fondo = COLOR_AZUL if activo else COLOR_BLANCO
    frente = COLOR_BLANCO if activo else COLOR_AZUL_OSCURO
    btn = tk.Button(panel, text=texto, command=comando, anchor="w", font=("Segoe UI", 11, "bold" if activo else "normal"),
                    bg=fondo, fg=frente, activebackground=COLOR_AZUL, activeforeground=COLOR_BLANCO,
                    bd=0, padx=18, pady=12, cursor="hand2")
    btn.pack(fill=tk.X, padx=14, pady=4)
    return btn
def tarjeta(parent, titulo, valor, subtitulo, color):
    frame = tk.Frame(parent, bg=COLOR_BLANCO, bd=1, relief="solid", highlightbackground=COLOR_GRIS, highlightthickness=1)
    head = tk.Frame(frame, bg=COLOR_BLANCO)
    head.pack(fill=tk.X, padx=15, pady=(14, 5))
    tk.Label(head, text="●", bg=COLOR_BLANCO, fg=color, font=("Segoe UI", 18, "bold")).pack(side=tk.LEFT)
    tk.Label(head, text=titulo, bg=COLOR_BLANCO, fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=8)
    tk.Label(frame, text=valor, bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 18, "bold")).pack(anchor="w", padx=18)
    tk.Label(frame, text=subtitulo, bg=COLOR_BLANCO, fg="#64748B", font=("Segoe UI", 9)).pack(anchor="w", padx=18, pady=(2, 14))
    return frame
def panel_titulo(parent, titulo):
    panel = tk.Frame(parent, bg=COLOR_BLANCO, bd=1, relief="solid", highlightbackground=COLOR_GRIS, highlightthickness=1)
    tk.Label(panel, text=titulo, bg=COLOR_BLANCO, fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=15, pady=12)
    return panel
def limpiar_tabla(tabla):
    for item in tabla.get_children():
        tabla.delete(item)
def mostrar_datos(tabla, columnas, resultados):
    limpiar_tabla(tabla)
    tabla["columns"] = columnas
    tabla["show"] = "headings"
    for col in columnas:
        tabla.heading(col, text=col)
        tabla.column(col, width=130)
    for fila in resultados:
        tabla.insert("", tk.END, values=[("" if dato is None else str(dato)) for dato in fila])
def ejecutar_consulta(tabla, consulta, parametros=None):
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        cur.execute(consulta, parametros or ())
        columnas = [col[0] for col in cur.description]
        mostrar_datos(tabla, columnas, cur.fetchall())
    except Exception as error:
        messagebox.showerror("Error", str(error))
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def crear_tabla(frame):
    cont = tk.Frame(frame, bg=COLOR_BLANCO)
    cont.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    tabla = ttk.Treeview(cont)
    sy = ttk.Scrollbar(cont, orient="vertical", command=tabla.yview)
    sx = ttk.Scrollbar(cont, orient="horizontal", command=tabla.xview)
    tabla.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)
    tabla.grid(row=0, column=0, sticky="nsew")
    sy.grid(row=0, column=1, sticky="ns")
    sx.grid(row=1, column=0, sticky="ew")
    cont.grid_rowconfigure(0, weight=1)
    cont.grid_columnconfigure(0, weight=1)
    return tabla
def consultar_resumen(tabla):
    ejecutar_consulta(tabla, "EXEC sp_resumen_general")
def consultar_meses(tabla):
    ejecutar_consulta(tabla, "EXEC sp_ventas_por_mes")
def consultar_categorias(tabla):
    ejecutar_consulta(tabla, "EXEC sp_ventas_por_categoria")
def buscar_ventas(tabla, entrada_producto, entrada_categoria):
    ejecutar_consulta(tabla, "EXEC sp_buscar_ventas ?, ?", (entrada_producto.get(), entrada_categoria.get()))
def consultar_catalogo(tabla, entrada_texto):
    texto = "%" + entrada_texto.get() + "%"
    consulta = """
        SELECT sku,categoria,producto,unidad,precio_venta,stock_actual,estado_stock,proveedor,ubicacion
        FROM Catalogo_Ampliado
        WHERE producto LIKE ? OR categoria LIKE ? OR sku LIKE ?
        ORDER BY categoria, producto
    """
    ejecutar_consulta(tabla, consulta, (texto, texto, texto))
def consultar_stock(tabla, combo_estado):
    ejecutar_consulta(tabla, "EXEC sp_consultar_stock ?", (combo_estado.get(),))
def obtener_skus_catalogo():
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        cur.execute("SELECT sku FROM Catalogo_Ampliado ORDER BY sku")
        return [fila.sku for fila in cur.fetchall()]
    except Exception as error:
        messagebox.showerror("Error", str(error))
        return []
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def obtener_datos_producto(sku):
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        cur.execute("""
            SELECT sku,categoria,producto,unidad,precio_venta,stock_actual
            FROM Catalogo_Ampliado
            WHERE sku=?
        """, sku)
        return cur.fetchone()
    except Exception as error:
        messagebox.showerror("Error", str(error))
        return None
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def generar_id_venta(cursor):
    cursor.execute("""
        SELECT ISNULL(MAX(TRY_CAST(SUBSTRING(id_venta,6,20) AS INT)),0)+1 AS numero
        FROM Ventas_2026
        WHERE id_venta LIKE 'VE26-%'
    """)
    return "VE26-" + str(cursor.fetchone().numero).zfill(6)
def calcular_estado_stock(stock_fisico, stock_minimo, punto_reorden):
    if stock_fisico <= stock_minimo:
        return "Crítico"
    if stock_fisico <= punto_reorden:
        return "Reordenar"
    if stock_fisico <= punto_reorden * 1.25:
        return "Vigilar"
    return "OK"
def copiar_formato_fila(hoja, origen, destino):
    for col in range(1, hoja.max_column + 1):
        c1 = hoja.cell(row=origen, column=col)
        c2 = hoja.cell(row=destino, column=col)
        if c1.has_style:
            c2._style = copy(c1._style)
        c2.font = copy(c1.font)
        c2.fill = copy(c1.fill)
        c2.border = copy(c1.border)
        c2.alignment = copy(c1.alignment)
        c2.number_format = c1.number_format
        c2.protection = copy(c1.protection)
    if hoja.row_dimensions[origen].height is not None:
        hoja.row_dimensions[destino].height = hoja.row_dimensions[origen].height
def poner_valor(hoja, mapa, fila, nombres, valor):
    for nombre in nombres:
        limpio = limpiar_texto(nombre)
        if limpio in mapa:
            hoja.cell(row=fila, column=mapa[limpio]).value = valor
            return True
    return False
def buscar_fila_sku(hoja, mapa, sku):
    if limpiar_texto("SKU") not in mapa:
        return None
    col = mapa[limpiar_texto("SKU")]
    for fila in range(2, hoja.max_row + 1):
        if str(hoja.cell(row=fila, column=col).value).strip() == str(sku).strip():
            return fila
    return None
def actualizar_excel_despues_venta(datos):
    ruta = obtener_ruta_excel_para_actualizar()
    if not ruta:
        raise Exception("No se selecciono ningun archivo Excel para actualizar.")
    libro = load_workbook(ruta)
    if "Ventas_2026" not in libro.sheetnames:
        raise Exception("El Excel no tiene la hoja Ventas_2026.")
    hoja = libro["Ventas_2026"]
    mapa = crear_mapa_columnas(hoja)
    fila = hoja.max_row + 1
    if hoja.max_row >= 2:
        copiar_formato_fila(hoja, hoja.max_row, fila)
    campos_venta = [
        (["ID_Venta", "ID Venta"], "id_venta"),
        (["Fecha"], "fecha"),
        (["Mes"], "mes"),
        (["Dia_Semana", "Día Semana", "Dia Semana"], "dia_semana"),
        (["Categoría", "Categoria"], "categoria"),
        (["SKU"], "sku"),
        (["Producto"], "producto"),
        (["Unidad"], "unidad"),
        (["Cantidad"], "cantidad"),
        (["Precio Unitario MXN", "Precio Unitario"], "precio_unitario"),
        (["Total MXN", "Total"], "total"),
        (["Metodo Pago", "Método Pago"], "metodo_pago"),
        (["Tipo Cliente"], "tipo_cliente"),
        (["Vendedor"], "vendedor"),
        (["Tipo Registro"], "tipo_registro"),
        (["Observacion", "Observación"], "observacion")
    ]
    for nombres, clave in campos_venta:
        poner_valor(hoja, mapa, fila, nombres, datos[clave])
    if "Control_Stock" in libro.sheetnames:
        hoja = libro["Control_Stock"]
        mapa = crear_mapa_columnas(hoja)
        fila = buscar_fila_sku(hoja, mapa, datos["sku"])
        if fila:
            for nombres, clave in [
                (["Salidas por Ventas", "Salidas Ventas"], "salidas"),
                (["Stock Calculado"], "stock_calculado"),
                (["Stock Físico Conteo", "Stock Fisico Conteo", "Stock Físico", "Stock Fisico"], "stock_fisico"),
                (["Diferencia"], "diferencia"),
                (["Estado"], "estado")
            ]:
                poner_valor(hoja, mapa, fila, nombres, datos[clave])
    if "Catalogo_Ampliado" in libro.sheetnames:
        hoja = libro["Catalogo_Ampliado"]
        mapa = crear_mapa_columnas(hoja)
        fila = buscar_fila_sku(hoja, mapa, datos["sku"])
        if fila:
            poner_valor(hoja, mapa, fila, ["Stock Actual"], datos["stock_fisico"])
            poner_valor(hoja, mapa, fila, ["Estado Stock"], datos["estado"])
    libro.save(ruta)
def registrar_venta_sql(sku, cantidad, fecha_venta, metodo_pago, tipo_cliente, vendedor, tipo_registro, observacion):
    con = None
    cur = None
    try:
        fecha = convertir_fecha_formulario(fecha_venta)
        if fecha is None:
            messagebox.showerror("Error", "La fecha debe tener formato AAAA-MM-DD. Ejemplo: 2026-07-01")
            return False
        con = conectar_base()
        cur = con.cursor()
        cur.execute("SELECT categoria,producto,unidad,precio_venta,stock_actual FROM Catalogo_Ampliado WHERE sku=?", sku)
        producto = cur.fetchone()
        if producto is None:
            messagebox.showerror("Error", "No existe ese SKU en el catalogo.")
            return False
        cur.execute("""
            SELECT stock_inicial,entradas_compras,salidas_ventas,merma_ajuste,stock_calculado,
                   stock_fisico,stock_minimo,punto_reorden,stock_maximo
            FROM Control_Stock
            WHERE sku=?
        """, sku)
        stock = cur.fetchone()
        if stock is None:
            messagebox.showerror("Error", "No existe ese SKU en control de stock.")
            return False
        meses = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
        dias = {0:"Lunes",1:"Martes",2:"Miercoles",3:"Jueves",4:"Viernes",5:"Sabado",6:"Domingo"}
        precio = float(producto.precio_venta)
        total = cantidad * precio
        nuevo_id = generar_id_venta(cur)
        nuevas_salidas = convertir_numero(stock.salidas_ventas) + cantidad
        nuevo_stock_calculado = convertir_numero(stock.stock_calculado) - cantidad
        nuevo_stock_fisico = convertir_numero(stock.stock_fisico) - cantidad
        diferencia = nuevo_stock_fisico - nuevo_stock_calculado
        estado = calcular_estado_stock(nuevo_stock_fisico, convertir_numero(stock.stock_minimo), convertir_numero(stock.punto_reorden))
        datos = {
            "id_venta": nuevo_id,
            "fecha": fecha,
            "mes": meses[fecha.month],
            "dia_semana": dias[fecha.weekday()],
            "categoria": producto.categoria,
            "sku": sku,
            "producto": producto.producto,
            "unidad": producto.unidad,
            "cantidad": cantidad,
            "precio_unitario": precio,
            "total": total,
            "metodo_pago": metodo_pago,
            "tipo_cliente": tipo_cliente,
            "vendedor": vendedor,
            "tipo_registro": tipo_registro,
            "observacion": observacion,
            "salidas": nuevas_salidas,
            "stock_calculado": nuevo_stock_calculado,
            "stock_fisico": nuevo_stock_fisico,
            "diferencia": diferencia,
            "estado": estado
        }
        cur.execute("""
            INSERT INTO Ventas_2026 (
                id_venta,fecha,mes,dia_semana,categoria,sku,producto,unidad,cantidad,
                precio_unitario,total_mxn,metodo_pago,tipo_cliente,vendedor,tipo_registro,observacion
            )
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        datos["id_venta"], datos["fecha"], datos["mes"], datos["dia_semana"], datos["categoria"],
        datos["sku"], datos["producto"], datos["unidad"], datos["cantidad"], datos["precio_unitario"],
        datos["total"], datos["metodo_pago"], datos["tipo_cliente"], datos["vendedor"], datos["tipo_registro"],
        datos["observacion"])
        cur.execute("""
            UPDATE Control_Stock
            SET salidas_ventas=?, stock_calculado=?, stock_fisico=?, diferencia=?, estado=?
            WHERE sku=?
        """, datos["salidas"], datos["stock_calculado"], datos["stock_fisico"], datos["diferencia"], datos["estado"], sku)
        cur.execute("UPDATE Catalogo_Ampliado SET stock_actual=?, estado_stock=? WHERE sku=?", datos["stock_fisico"], datos["estado"], sku)
        actualizar_excel_despues_venta(datos)
        con.commit()
        messagebox.showinfo(
            "Venta registrada",
            "Venta registrada correctamente en SQL Server y Excel.\n\n"
            + "ID venta: " + nuevo_id + "\n"
            + "Fecha: " + str(fecha) + "\n"
            + "Producto: " + str(producto.producto) + "\n"
            + "Cantidad: " + str(cantidad) + "\n"
            + "Total: $" + str(round(total, 2)) + "\n"
            + "Nuevo stock: " + str(round(nuevo_stock_fisico, 2)) + "\n"
            + "Estado stock: " + estado
        )
        return True
    except PermissionError:
        if con:
            con.rollback()
        messagebox.showerror("Excel abierto", "Cierra el archivo Excel y vuelve a registrar la venta.")
        return False
    except Exception as error:
        if con:
            con.rollback()
        messagebox.showerror("Error al registrar venta", str(error))
        return False
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def crear_formulario_nueva_venta(parent):
    panel = tk.Frame(parent, bg=COLOR_BLANCO, bd=1, relief="solid", highlightbackground=COLOR_GRIS, highlightthickness=1)
    panel.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    head = tk.Frame(panel, bg=COLOR_BLANCO)
    head.pack(fill=tk.X, padx=25, pady=(20, 10))
    tk.Label(head, text="Nueva venta", bg=COLOR_BLANCO, fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 18, "bold")).pack(side=tk.LEFT)
    tk.Label(head, text="Registra una venta y actualiza SQL Server + Excel", bg=COLOR_BLANCO, fg="#64748B", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=15)
    cuerpo = tk.Frame(panel, bg=COLOR_BLANCO)
    cuerpo.pack(fill=tk.BOTH, expand=True, padx=25, pady=10)
    form = tk.Frame(cuerpo, bg=COLOR_BLANCO)
    form.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 20))
    resumen = tk.Frame(cuerpo, bg=COLOR_AZUL_CLARO, bd=1, relief="solid", highlightbackground=COLOR_GRIS, highlightthickness=1, width=320)
    resumen.pack(side=tk.RIGHT, fill=tk.Y)
    resumen.pack_propagate(False)
    def label(texto, fila, col):
        tk.Label(form, text=texto, bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).grid(row=fila, column=col, sticky="w", padx=5, pady=(8, 2))
    def entry(fila, col, ancho=32, readonly=False):
        caja = tk.Entry(form, width=ancho, font=("Segoe UI", 10), bd=1, relief="solid")
        caja.grid(row=fila, column=col, sticky="w", padx=5, pady=(2, 8), ipady=6)
        if readonly:
            caja.config(state="readonly")
        return caja
    label("Fecha", 0, 0)
    entrada_fecha = entry(1, 0)
    entrada_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))
    label("SKU", 0, 1)
    combo_sku = ttk.Combobox(form, values=obtener_skus_catalogo(), width=30)
    combo_sku.grid(row=1, column=1, sticky="w", padx=5, pady=(2, 8), ipady=4)
    label("Producto", 2, 0)
    entrada_producto = entry(3, 0, 70, True)
    label("Categoría", 2, 1)
    entrada_categoria = entry(3, 1, 32, True)
    label("Unidad", 4, 0)
    entrada_unidad = entry(5, 0, 32, True)
    label("Precio unitario", 4, 1)
    entrada_precio = entry(5, 1, 32, True)
    label("Stock actual", 6, 0)
    entrada_stock = entry(7, 0, 32, True)
    label("Cantidad", 6, 1)
    entrada_cantidad = entry(7, 1, 32)
    label("Método de pago", 8, 0)
    combo_pago = ttk.Combobox(form, values=["Efectivo", "Transferencia", "Tarjeta", "Credito local"], width=30)
    combo_pago.grid(row=9, column=0, sticky="w", padx=5, pady=(2, 8), ipady=4)
    combo_pago.set("Efectivo")
    label("Tipo cliente", 8, 1)
    combo_cliente = ttk.Combobox(form, values=["Publico general", "Contratista local", "Agricultor/Jardinero", "Reparacion domestica", "Cliente frecuente"], width=30)
    combo_cliente.grid(row=9, column=1, sticky="w", padx=5, pady=(2, 8), ipady=4)
    combo_cliente.set("Publico general")
    label("Vendedor", 10, 0)
    combo_vendedor = ttk.Combobox(form, values=["Mostrador 1", "Mostrador 2", "Caja/Bodega"], width=30)
    combo_vendedor.grid(row=11, column=0, sticky="w", padx=5, pady=(2, 8), ipady=4)
    combo_vendedor.set("Mostrador 1")
    label("Tipo registro", 10, 1)
    combo_tipo = ttk.Combobox(form, values=["Completo", "Completo verificado", "Ajuste menor", "Recuperado de ticket"], width=30)
    combo_tipo.grid(row=11, column=1, sticky="w", padx=5, pady=(2, 8), ipady=4)
    combo_tipo.set("Completo")
    label("Observación", 12, 0)
    entrada_observacion = entry(13, 0, 70)
    tk.Label(resumen, text="Resumen de venta", bg=COLOR_AZUL_CLARO, fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 15, "bold")).pack(anchor="w", padx=20, pady=(22, 10))
    resumen_producto = tk.Label(resumen, text="Producto: -", bg=COLOR_AZUL_CLARO, fg=COLOR_TEXTO, wraplength=270, justify="left", font=("Segoe UI", 10))
    resumen_producto.pack(anchor="w", padx=20, pady=5)
    resumen_cantidad = tk.Label(resumen, text="Cantidad: 0", bg=COLOR_AZUL_CLARO, fg=COLOR_TEXTO, font=("Segoe UI", 10))
    resumen_cantidad.pack(anchor="w", padx=20, pady=5)
    resumen_precio = tk.Label(resumen, text="Precio unitario: $0.00", bg=COLOR_AZUL_CLARO, fg=COLOR_TEXTO, font=("Segoe UI", 10))
    resumen_precio.pack(anchor="w", padx=20, pady=5)
    tk.Frame(resumen, bg=COLOR_GRIS, height=1).pack(fill=tk.X, padx=20, pady=15)
    resumen_total = tk.Label(resumen, text="TOTAL: $0.00", bg=COLOR_AZUL_CLARO, fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 18, "bold"))
    resumen_total.pack(anchor="w", padx=20, pady=10)
    def readonly(caja, texto):
        caja.config(state="normal")
        caja.delete(0, tk.END)
        caja.insert(0, texto)
        caja.config(state="readonly")
    def cargar_producto(evento=None):
        sku = combo_sku.get()
        if not sku:
            return
        datos = obtener_datos_producto(sku)
        if datos is None:
            return
        readonly(entrada_producto, datos.producto)
        readonly(entrada_categoria, datos.categoria)
        readonly(entrada_unidad, datos.unidad)
        readonly(entrada_precio, datos.precio_venta)
        readonly(entrada_stock, datos.stock_actual)
        resumen_producto.config(text="Producto: " + str(datos.producto))
        calcular_total()
    def calcular_total(evento=None):
        try:
            cantidad = float(entrada_cantidad.get())
            precio = float(entrada_precio.get())
            total = cantidad * precio
            resumen_cantidad.config(text="Cantidad: " + str(cantidad))
            resumen_precio.config(text="Precio unitario: $" + str(round(precio, 2)))
            resumen_total.config(text="TOTAL: $" + str(round(total, 2)))
        except:
            resumen_cantidad.config(text="Cantidad: 0")
            resumen_precio.config(text="Precio unitario: $0.00")
            resumen_total.config(text="TOTAL: $0.00")
    def limpiar_formulario():
        entrada_fecha.delete(0, tk.END)
        entrada_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))
        combo_sku.set("")
        for caja in [entrada_producto, entrada_categoria, entrada_unidad, entrada_precio, entrada_stock]:
            readonly(caja, "")
        entrada_cantidad.delete(0, tk.END)
        combo_pago.set("Efectivo")
        combo_cliente.set("Publico general")
        combo_vendedor.set("Mostrador 1")
        combo_tipo.set("Completo")
        entrada_observacion.delete(0, tk.END)
        resumen_producto.config(text="Producto: -")
        resumen_cantidad.config(text="Cantidad: 0")
        resumen_precio.config(text="Precio unitario: $0.00")
        resumen_total.config(text="TOTAL: $0.00")
    def registrar_desde_formulario():
        fecha = entrada_fecha.get()
        sku = combo_sku.get()
        cantidad_txt = entrada_cantidad.get()
        if convertir_fecha_formulario(fecha) is None:
            messagebox.showerror("Error", "La fecha debe tener formato AAAA-MM-DD. Ejemplo: 2026-07-01")
            return
        if not sku:
            messagebox.showwarning("Faltan datos", "Selecciona un SKU.")
            return
        if not cantidad_txt:
            messagebox.showwarning("Faltan datos", "Escribe la cantidad.")
            return
        try:
            cantidad = float(cantidad_txt)
        except:
            messagebox.showerror("Error", "La cantidad debe ser numerica.")
            return
        if cantidad <= 0:
            messagebox.showerror("Error", "La cantidad debe ser mayor a cero.")
            return
        try:
            stock = float(entrada_stock.get())
        except:
            stock = 0
        if cantidad > stock:
            ok = messagebox.askyesno("Stock insuficiente", f"La cantidad supera el stock actual.\n\nStock actual: {stock}\nCantidad venta: {cantidad}\n\n¿Deseas registrar la venta de todos modos?")
            if not ok:
                return
        registrado = registrar_venta_sql(
            sku, cantidad, fecha, combo_pago.get(), combo_cliente.get(),
            combo_vendedor.get(), combo_tipo.get(), entrada_observacion.get()
        )
        if registrado:
            limpiar_formulario()
    combo_sku.bind("<<ComboboxSelected>>", cargar_producto)
    entrada_cantidad.bind("<KeyRelease>", calcular_total)
    tk.Button(resumen, text="REGISTRAR VENTA", bg=COLOR_AZUL, fg=COLOR_BLANCO,
              activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
              font=("Segoe UI", 13, "bold"), bd=0, cursor="hand2",
              command=registrar_desde_formulario).pack(fill=tk.X, padx=20, pady=(25, 10), ipady=10)
    tk.Button(resumen, text="LIMPIAR", bg=COLOR_GRIS, fg=COLOR_TEXTO,
              activebackground="#CBD5E1", activeforeground=COLOR_TEXTO,
              font=("Segoe UI", 11, "bold"), bd=0, cursor="hand2",
              command=limpiar_formulario).pack(fill=tk.X, padx=20, pady=(0, 10), ipady=8)
def verificar_tabla_usuarios(cursor):
    cursor.execute("SELECT OBJECT_ID('Usuarios','U') AS id_tabla")
    fila = cursor.fetchone()
    return fila is not None and fila.id_tabla is not None
def asegurar_usuario_admin():
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        if not verificar_tabla_usuarios(cur):
            return False, "No existe la tabla Usuarios. Ejecuta primero la opcion 1."
        cur.execute("SELECT COUNT(*) AS total FROM Usuarios WHERE usuario=?", "admin")
        if cur.fetchone().total == 0:
            cur.execute("""
                DECLARE @salt VARBINARY(16)
                DECLARE @hash VARBINARY(32)
                SET @salt = CRYPT_GEN_RANDOM(16)
                SET @hash = HASHBYTES('SHA2_256', @salt + CONVERT(VARBINARY(MAX), ?))
                INSERT INTO Usuarios(usuario,password_hash,password_salt,rol,activo)
                VALUES(?,@hash,@salt,?,1)
            """, "1234", "admin", "Administrador")
            con.commit()
        return True, ""
    except Exception as error:
        if con:
            try:
                con.rollback()
            except:
                pass
        return False, str(error)
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def login(ventana, entrada_usuario, entrada_contrasena):
    usuario = entrada_usuario.get().strip()
    contrasena = entrada_contrasena.get().strip()
    if not usuario or not contrasena:
        messagebox.showwarning("Faltan datos", "Escribe usuario y contraseña.")
        return
    ok, msg = asegurar_usuario_admin()
    if not ok:
        messagebox.showerror("Error de preparacion", msg + "\n\nEjecuta 1. Preparar SQL Server y luego 3. Importar Excel.")
        return
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        cur.execute("EXEC sp_login ?, ?", usuario, contrasena)
        fila = cur.fetchone()
        if fila is None:
            messagebox.showerror("Error", "Usuario o contraseña incorrectos.\n\nUsuario: admin\nContraseña: 1234")
        else:
            ventana.destroy()
            abrir_sistema()
    except Exception as error:
        messagebox.showerror("Error al iniciar sesion", str(error))
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def abrir_login():
    configurar_estilos()
    v = tk.Tk()
    v.title("Ferreteria EL VERGEL S.A. de C.V. - Login")
    v.geometry("430x430")
    v.configure(bg=COLOR_FONDO)
    v.resizable(False, False)
    barra = tk.Frame(v, bg=COLOR_AZUL_OSCURO, height=60)
    barra.pack(fill=tk.X)
    tk.Label(barra, text="Ferreteria EL VERGEL S.A. de C.V.", bg=COLOR_AZUL_OSCURO,
             fg=COLOR_BLANCO, font=("Segoe UI", 14, "bold")).pack(pady=15)
    card = tk.Frame(v, bg=COLOR_BLANCO, bd=1, relief="solid")
    card.pack(padx=35, pady=30, fill=tk.BOTH, expand=True)
    tk.Label(card, text="Inicio de sesión", font=("Segoe UI", 17, "bold"),
             bg=COLOR_BLANCO, fg=COLOR_AZUL_OSCURO).pack(pady=(20, 15))
    tk.Label(card, text="Usuario", bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=40)
    usuario = tk.Entry(card, font=("Segoe UI", 11), bd=1, relief="solid")
    usuario.pack(fill=tk.X, padx=40, pady=(4, 12), ipady=6)
    usuario.insert(0, "admin")
    tk.Label(card, text="Contraseña", bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=40)
    clave = tk.Entry(card, show="*", font=("Segoe UI", 11), bd=1, relief="solid")
    clave.pack(fill=tk.X, padx=40, pady=(4, 18), ipady=6)
    clave.insert(0, "1234")
    tk.Button(card, text="ENTRAR", bg=COLOR_AZUL, fg=COLOR_BLANCO,
              activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
              font=("Segoe UI", 12, "bold"), bd=0, cursor="hand2",
              command=lambda: login(v, usuario, clave)).pack(fill=tk.X, padx=40, pady=20, ipady=8)
    tk.Label(card, text="Usuario: admin   |   Contraseña: 1234",
             bg=COLOR_BLANCO, fg="#64748B", font=("Segoe UI", 9)).pack(pady=(0, 8))
    v.bind("<Return>", lambda e: login(v, usuario, clave))
    clave.focus_set()
    v.mainloop()
def abrir_sistema():
    configurar_estilos()
    v = tk.Tk()
    v.title("Ferreteria EL VERGEL S.A. de C.V.")
    v.geometry("1350x760")
    v.configure(bg=COLOR_FONDO)
    v.minsize(1150, 680)
    top = tk.Frame(v, bg=COLOR_AZUL_OSCURO, height=58)
    top.pack(fill=tk.X)
    top.pack_propagate(False)
    tk.Label(top, text="Ferreteria EL VERGEL S.A. de C.V.", bg=COLOR_AZUL_OSCURO,
             fg=COLOR_BLANCO, font=("Segoe UI", 18, "bold")).pack(side=tk.LEFT, padx=25)
    tk.Label(top, text="Sistema de ventas, catálogo y control de stock", bg=COLOR_AZUL_OSCURO,
             fg="#CFE3FF", font=("Segoe UI", 10)).pack(side=tk.RIGHT, padx=25)
    root = tk.Frame(v, bg=COLOR_FONDO)
    root.pack(fill=tk.BOTH, expand=True)
    side = tk.Frame(root, bg=COLOR_BLANCO, width=285)
    side.pack(side=tk.LEFT, fill=tk.Y)
    side.pack_propagate(False)
    logo = cargar_logo_tk()
    if logo:
        v.logo_empresa = logo
        tk.Label(side, image=logo, bg=COLOR_BLANCO).pack(padx=12, pady=(18, 10))
    else:
        tk.Label(side, text="FERRETERIA\nEL VERGEL", bg=COLOR_BLANCO,
                 fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 20, "bold"),
                 justify="center").pack(pady=(25, 18))
    tk.Label(side, text="Punto de venta", bg=COLOR_BLANCO, fg="#64748B",
             font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=22, pady=(5, 8))
    area = tk.Frame(root, bg=COLOR_FONDO)
    area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    encabezado = tk.Frame(area, bg=COLOR_FONDO, height=80)
    encabezado.pack(fill=tk.X)
    encabezado.pack_propagate(False)
    tk.Label(encabezado, text="Bienvenido, Administrador", bg=COLOR_FONDO,
             fg=COLOR_AZUL_OSCURO, font=("Segoe UI", 16, "bold")).pack(anchor="w", padx=25, pady=(18, 0))
    tk.Label(encabezado, text="Control general de ventas, catálogo e inventario de Ferreteria EL VERGEL.",
             bg=COLOR_FONDO, fg="#64748B", font=("Segoe UI", 10)).pack(anchor="w", padx=25)
    contenido = tk.Frame(area, bg=COLOR_FONDO)
    contenido.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    def limpiar():
        for widget in contenido.winfo_children():
            widget.destroy()
    def pantalla_resumen():
        limpiar()
        cards = tk.Frame(contenido, bg=COLOR_FONDO)
        cards.pack(fill=tk.X, pady=(0, 15))
        for datos in [
            ("Resumen general", "SQL Server", "Consulta los indicadores cargados", COLOR_AZUL),
            ("Ventas por mes", "2026", "Agrupación mensual de ventas", COLOR_VERDE),
            ("Ventas por categoría", "Catálogo", "Comparativo por departamento", COLOR_NARANJA)
        ]:
            tarjeta(cards, *datos).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
        botones = tk.Frame(contenido, bg=COLOR_BLANCO)
        botones.pack(fill=tk.X, pady=(0, 10))
        tabla = crear_tabla(contenido)
        acciones = [
            ("Resumen general", lambda: consultar_resumen(tabla)),
            ("Ventas por mes", lambda: consultar_meses(tabla)),
            ("Ventas por categoría", lambda: consultar_categorias(tabla))
        ]
        for texto, comando in acciones:
            tk.Button(botones, text=texto, bg=COLOR_AZUL, fg=COLOR_BLANCO,
                      activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                      font=("Segoe UI", 10, "bold"), bd=0, padx=16, pady=9,
                      cursor="hand2", command=comando).pack(side=tk.LEFT, padx=8, pady=10)
        consultar_resumen(tabla)
    def pantalla_ventas():
        limpiar()
        panel = panel_titulo(contenido, "Ventas registradas")
        panel.pack(fill=tk.X, pady=(0, 10))
        filtros = tk.Frame(panel, bg=COLOR_BLANCO)
        filtros.pack(fill=tk.X, padx=15, pady=(0, 15))
        tk.Label(filtros, text="Producto", bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        prod = tk.Entry(filtros, width=25, font=("Segoe UI", 10))
        prod.pack(side=tk.LEFT, padx=5, ipady=5)
        tk.Label(filtros, text="Categoría", bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=(15, 5))
        cat = tk.Entry(filtros, width=25, font=("Segoe UI", 10))
        cat.pack(side=tk.LEFT, padx=5, ipady=5)
        tabla = crear_tabla(contenido)
        tk.Button(filtros, text="Buscar", bg=COLOR_AZUL, fg=COLOR_BLANCO,
                  activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=18, pady=7,
                  cursor="hand2", command=lambda: buscar_ventas(tabla, prod, cat)).pack(side=tk.LEFT, padx=12)
    def pantalla_catalogo():
        limpiar()
        panel = panel_titulo(contenido, "Catálogo de productos")
        panel.pack(fill=tk.X, pady=(0, 10))
        filtros = tk.Frame(panel, bg=COLOR_BLANCO)
        filtros.pack(fill=tk.X, padx=15, pady=(0, 15))
        tk.Label(filtros, text="Buscar producto o SKU", bg=COLOR_BLANCO, fg=COLOR_TEXTO,
                 font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        entrada = tk.Entry(filtros, width=38, font=("Segoe UI", 10))
        entrada.pack(side=tk.LEFT, padx=5, ipady=5)
        tabla = crear_tabla(contenido)
        tk.Button(filtros, text="Buscar", bg=COLOR_AZUL, fg=COLOR_BLANCO,
                  activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=18, pady=7,
                  cursor="hand2", command=lambda: consultar_catalogo(tabla, entrada)).pack(side=tk.LEFT, padx=12)
    def pantalla_stock():
        limpiar()
        panel = panel_titulo(contenido, "Control de stock")
        panel.pack(fill=tk.X, pady=(0, 10))
        filtros = tk.Frame(panel, bg=COLOR_BLANCO)
        filtros.pack(fill=tk.X, padx=15, pady=(0, 15))
        tk.Label(filtros, text="Estado", bg=COLOR_BLANCO, fg=COLOR_TEXTO, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT)
        combo = ttk.Combobox(filtros, values=["", "Crítico", "Reordenar", "Vigilar", "OK"], width=22)
        combo.pack(side=tk.LEFT, padx=5)
        tabla = crear_tabla(contenido)
        tk.Button(filtros, text="Consultar stock", bg=COLOR_AZUL, fg=COLOR_BLANCO,
                  activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=18, pady=7,
                  cursor="hand2", command=lambda: consultar_stock(tabla, combo)).pack(side=tk.LEFT, padx=12)
    def pantalla_historial():
        limpiar()
        panel = panel_titulo(contenido, "Historial de ventas 2024 y 2025")
        panel.pack(fill=tk.X, pady=(0, 10))
        barra = tk.Frame(panel, bg=COLOR_BLANCO)
        barra.pack(fill=tk.X, padx=15, pady=(0, 15))
        tk.Label(barra, text="Consulta opcional de datasets historicos. No usa SQL Server.",
                 bg=COLOR_BLANCO, fg="#64748B", font=("Segoe UI", 10)).pack(side=tk.LEFT, padx=(0, 12))
        tabla = crear_tabla(contenido)
        tk.Button(barra, text="Ver 2024", bg=COLOR_AZUL, fg=COLOR_BLANCO,
                  activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=14, pady=7,
                  cursor="hand2", command=lambda: cargar_dataset_en_tabla(tabla, DATASET_VENTAS_2024)).pack(side=tk.LEFT, padx=4)
        tk.Button(barra, text="Ver 2025", bg=COLOR_AZUL, fg=COLOR_BLANCO,
                  activebackground=COLOR_AZUL_OSCURO, activeforeground=COLOR_BLANCO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=14, pady=7,
                  cursor="hand2", command=lambda: cargar_dataset_en_tabla(tabla, DATASET_VENTAS_2025)).pack(side=tk.LEFT, padx=4)
        tk.Button(barra, text="Abrir Excel 2024", bg=COLOR_GRIS, fg=COLOR_TEXTO,
                  activebackground="#CBD5E1", activeforeground=COLOR_TEXTO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=14, pady=7,
                  cursor="hand2", command=abrir_dataset_2024).pack(side=tk.LEFT, padx=4)
        tk.Button(barra, text="Abrir Excel 2025", bg=COLOR_GRIS, fg=COLOR_TEXTO,
                  activebackground="#CBD5E1", activeforeground=COLOR_TEXTO,
                  font=("Segoe UI", 10, "bold"), bd=0, padx=14, pady=7,
                  cursor="hand2", command=abrir_dataset_2025).pack(side=tk.LEFT, padx=4)
    def pantalla_nueva_venta():
        limpiar()
        cont = tk.Frame(contenido, bg=COLOR_FONDO)
        cont.pack(fill=tk.BOTH, expand=True)
        crear_formulario_nueva_venta(cont)
    for texto, comando, activo in [
        ("Resumen", pantalla_resumen, True),
        ("Ventas", pantalla_ventas, False),
        ("Catálogo", pantalla_catalogo, False),
        ("Stock", pantalla_stock, False),
        ("Nueva venta", pantalla_nueva_venta, False),
        ("Reportes", abrir_power_bi, False),
        ("Historial de ventas", pantalla_historial, False)
    ]:
        boton_menu(side, texto, comando, activo)
    tk.Frame(side, bg=COLOR_GRIS, height=1).pack(fill=tk.X, padx=18, pady=22)
    boton_menu(side, "Configuración", pantalla_resumen)
    boton_menu(side, "Cerrar sesión", v.destroy)
    pantalla_resumen()
    v.mainloop()
def probar_conexion():
    con = None
    cur = None
    try:
        con = conectar_base()
        cur = con.cursor()
        cur.execute("SELECT DB_NAME() AS base_actual")
        print("\nConexion correcta")
        print("Base conectada:", cur.fetchone().base_actual)
    except Exception as error:
        print("\nError de conexion:")
        print(error)
    finally:
        if cur:
            cur.close()
        if con:
            con.close()
def menu():
    while True:
        print("\n===================================")
        print(" SISTEMA FERRETERIA EL VERGEL")
        print("===================================")
        print("1. Preparar SQL Server")
        print("2. Probar conexion")
        print("3. Importar Excel a SQL Server")
        print("4. Abrir interfaz")
        print("5. Salir")
        print("===================================")
        opcion = input("Elige una opcion: ")
        if opcion == "1":
            preparar_sql_server()
        elif opcion == "2":
            probar_conexion()
        elif opcion == "3":
            importar_excel()
        elif opcion == "4":
            abrir_login()
        elif opcion == "5":
            print("Saliendo...")
            break
        else:
            print("Opcion no valida.")
menu()
